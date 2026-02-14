#!/usr/bin/env python3
"""
AUDIT GLOBAL SYSTÈME IAPF
==========================

Audit complet structuré du système :
- Repo 1 : OCR Intelligent (Cloud Run)
- Repo 2 : CRM Box Magique (à analyser)
- Google Sheets : HUB + BOX2026
- MCP Cockpit : Gouvernance

Mode : PROPOSAL-FIRST (lecture seule, aucune modification)
"""

import json
import os
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl

# Constants
WEBAPP_DIR = Path("/home/user/webapp")
UPLOADED_FILES = Path("/home/user/uploaded_files")
HUB_FILE = UPLOADED_FILES / "IAPF_MEMORY_HUB_V1 (13).xlsx"
BOX_FILE = UPLOADED_FILES / "BOX2026 IAPF Cyril MARTINS (2).xlsx"


class AuditReport:
    """Gestionnaire de rapport d'audit structuré"""
    
    def __init__(self):
        self.timestamp = datetime.now().isoformat()
        self.sections = {}
        self.proposals = []
        self.corrections = []
        self.risks = []
        self.conflicts = []
        
    def add_section(self, name: str, data: dict):
        """Ajoute une section au rapport"""
        self.sections[name] = data
        
    def add_proposal(self, category: str, description: str, priority: str = "medium"):
        """Ajoute une proposition"""
        self.proposals.append({
            "category": category,
            "description": description,
            "priority": priority,
            "timestamp": datetime.now().isoformat()
        })
        
    def add_correction(self, area: str, issue: str, solution: str, authorized: bool = True):
        """Ajoute une correction ciblée"""
        self.corrections.append({
            "area": area,
            "issue": issue,
            "solution": solution,
            "authorized": authorized,
            "timestamp": datetime.now().isoformat()
        })
        
    def add_risk(self, level: str, description: str, impact: str):
        """Ajoute un risque identifié"""
        self.risks.append({
            "level": level,
            "description": description,
            "impact": impact,
            "timestamp": datetime.now().isoformat()
        })
        
    def add_conflict(self, location: str, description: str, severity: str):
        """Ajoute un conflit détecté"""
        self.conflicts.append({
            "location": location,
            "description": description,
            "severity": severity,
            "timestamp": datetime.now().isoformat()
        })
        
    def to_dict(self) -> dict:
        """Exporte le rapport en dictionnaire"""
        return {
            "meta": {
                "timestamp": self.timestamp,
                "version": "1.0.0",
                "mode": "PROPOSAL_FIRST"
            },
            "sections": self.sections,
            "proposals": self.proposals,
            "corrections": self.corrections,
            "risks": self.risks,
            "conflicts": self.conflicts,
            "stats": {
                "sections_count": len(self.sections),
                "proposals_count": len(self.proposals),
                "corrections_count": len(self.corrections),
                "risks_count": len(self.risks),
                "conflicts_count": len(self.conflicts)
            }
        }
        
    def save(self, output_path: Path):
        """Sauvegarde le rapport en JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)
        print(f"✅ Rapport sauvegardé : {output_path}")


class OCRAudit:
    """Audit du Repo 1 - OCR Intelligent"""
    
    def __init__(self, webapp_dir: Path):
        self.webapp_dir = webapp_dir
        self.results = {}
        
    def run(self) -> dict:
        """Exécute l'audit OCR complet"""
        print("\n" + "="*80)
        print("🔍 AUDIT REPO 1 - OCR INTELLIGENT")
        print("="*80)
        
        # 1. Vérifier architecture fichiers
        self._audit_structure()
        
        # 2. Vérifier pipeline multi-niveaux
        self._audit_pipeline()
        
        # 3. Vérifier extraction et scoring
        self._audit_extraction()
        
        # 4. Vérifier gouvernance READ-ONLY
        self._audit_governance()
        
        # 5. Vérifier intégration Cloud Run
        self._audit_cloudrun()
        
        return self.results
        
    def _audit_structure(self):
        """Audit de la structure des fichiers"""
        print("\n📂 Audit structure fichiers...")
        
        required_files = [
            "main.py",
            "ocr_engine.py",
            "requirements.txt",
            "Dockerfile",
            "config/config.yaml"
        ]
        
        required_dirs = [
            "levels",
            "memory",
            "connectors",
            "utils",
            "mcp_cockpit"
        ]
        
        structure = {
            "files_present": [],
            "files_missing": [],
            "dirs_present": [],
            "dirs_missing": []
        }
        
        for file in required_files:
            path = self.webapp_dir / file
            if path.exists():
                structure["files_present"].append(file)
            else:
                structure["files_missing"].append(file)
                
        for dir in required_dirs:
            path = self.webapp_dir / dir
            if path.is_dir():
                structure["dirs_present"].append(dir)
            else:
                structure["dirs_missing"].append(dir)
                
        self.results["structure"] = structure
        
        print(f"  ✅ Fichiers présents : {len(structure['files_present'])}/{len(required_files)}")
        print(f"  ✅ Dossiers présents : {len(structure['dirs_present'])}/{len(required_dirs)}")
        
        if structure["files_missing"]:
            print(f"  ⚠️  Fichiers manquants : {structure['files_missing']}")
        if structure["dirs_missing"]:
            print(f"  ⚠️  Dossiers manquants : {structure['dirs_missing']}")
            
    def _audit_pipeline(self):
        """Audit du pipeline OCR multi-niveaux"""
        print("\n🔄 Audit pipeline OCR...")
        
        pipeline = {
            "levels": {},
            "flow": "sequential",
            "memory_integration": False
        }
        
        # Vérifier les 3 niveaux
        for level in [1, 2, 3]:
            level_file = self.webapp_dir / f"levels/ocr_level{level}.py"
            if level_file.exists():
                size = level_file.stat().st_size
                with open(level_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    pipeline["levels"][f"level{level}"] = {
                        "exists": True,
                        "size": size,
                        "has_process_method": "def process(" in content,
                        "has_confidence_scoring": "confidence" in content.lower(),
                        "has_field_extraction": "fields" in content.lower()
                    }
                print(f"  ✅ OCR Level {level} détecté ({size} bytes)")
            else:
                pipeline["levels"][f"level{level}"] = {"exists": False}
                print(f"  ❌ OCR Level {level} manquant")
                
        # Vérifier mémoire
        memory_file = self.webapp_dir / "memory/ai_memory.py"
        if memory_file.exists():
            pipeline["memory_integration"] = True
            print(f"  ✅ AI Memory intégré")
        else:
            print(f"  ❌ AI Memory manquant")
            
        self.results["pipeline"] = pipeline
        
    def _audit_extraction(self):
        """Audit de l'extraction et scoring"""
        print("\n🎯 Audit extraction & scoring...")
        
        extraction = {
            "document_types": [],
            "field_extraction": {},
            "scoring_system": False,
            "fallback_vision": False
        }
        
        # Analyser ocr_engine.py
        engine_file = self.webapp_dir / "ocr_engine.py"
        if engine_file.exists():
            with open(engine_file, 'r', encoding='utf-8') as f:
                content = f.read()
                
                # Détection type de document
                if "document_type" in content.lower():
                    extraction["has_type_detection"] = True
                    print("  ✅ Détection type document présente")
                    
                # Extraction HT/TVA/TTC
                for field in ["ht", "tva", "ttc"]:
                    if field in content.lower():
                        extraction["field_extraction"][field] = True
                        
                # Système de scoring
                if "confidence" in content.lower() and "score" in content.lower():
                    extraction["scoring_system"] = True
                    print("  ✅ Système de scoring confiance détecté")
                    
                # Séparation entreprise/client
                if "entreprise_source" in content:
                    extraction["source_separation"] = True
                    print("  ✅ Séparation entreprise_source/client présente")
                    
        print(f"  ✅ Champs extraits identifiés : {list(extraction['field_extraction'].keys())}")
        
        self.results["extraction"] = extraction
        
    def _audit_governance(self):
        """Audit de la gouvernance READ-ONLY"""
        print("\n🔒 Audit gouvernance READ-ONLY...")
        
        governance = {
            "read_only_enforced": False,
            "no_sheets_write": False,
            "no_drive_write": False,
            "json_output_only": False
        }
        
        # Vérifier dans ocr_engine.py
        engine_file = self.webapp_dir / "ocr_engine.py"
        if engine_file.exists():
            with open(engine_file, 'r', encoding='utf-8') as f:
                content = f.read()
                
                if "OCR_READ_ONLY" in content:
                    governance["read_only_enforced"] = True
                    print("  ✅ OCR_READ_ONLY flag détecté")
                    
                if "[GOV]" in content:
                    governance["has_governance_markers"] = True
                    print("  ✅ Marqueurs gouvernance présents")
                    
                if "sheets_connector = None" in content:
                    governance["no_sheets_write"] = True
                    print("  ✅ Sheets connector désactivé (READ-ONLY)")
                    
                if "returns JSON" in content or "JSON result" in content:
                    governance["json_output_only"] = True
                    print("  ✅ Output JSON uniquement confirmé")
                    
        self.results["governance"] = governance
        
    def _audit_cloudrun(self):
        """Audit intégration Cloud Run"""
        print("\n☁️  Audit intégration Cloud Run...")
        
        cloudrun = {
            "dockerfile_exists": False,
            "fastapi_detected": False,
            "health_endpoint": False,
            "ocr_endpoint": False
        }
        
        # Vérifier Dockerfile
        dockerfile = self.webapp_dir / "Dockerfile"
        if dockerfile.exists():
            cloudrun["dockerfile_exists"] = True
            print("  ✅ Dockerfile présent")
            
        # Vérifier main.py (FastAPI)
        main_file = self.webapp_dir / "main.py"
        if main_file.exists():
            with open(main_file, 'r', encoding='utf-8') as f:
                content = f.read()
                
                if "fastapi" in content.lower():
                    cloudrun["fastapi_detected"] = True
                    print("  ✅ FastAPI détecté")
                    
                if "@app.get(\"/health\")" in content:
                    cloudrun["health_endpoint"] = True
                    print("  ✅ Endpoint /health présent")
                    
                if "@app.post(\"/ocr\")" in content:
                    cloudrun["ocr_endpoint"] = True
                    print("  ✅ Endpoint /ocr présent")
                    
        self.results["cloudrun"] = cloudrun


class HUBAudit:
    """Audit du HUB ORION (IAPF_MEMORY_HUB_V1)"""
    
    def __init__(self, hub_file: Path):
        self.hub_file = hub_file
        self.results = {}
        
    def run(self) -> dict:
        """Exécute l'audit HUB complet"""
        print("\n" + "="*80)
        print("🔍 AUDIT HUB ORION (MEMORY_HUB)")
        print("="*80)
        
        if not self.hub_file.exists():
            print(f"  ❌ Fichier HUB non trouvé : {self.hub_file}")
            self.results["error"] = "HUB file not found"
            return self.results
            
        # Charger le workbook
        wb = openpyxl.load_workbook(self.hub_file, read_only=True)
        
        # 1. Vérifier onglets critiques
        self._audit_sheets(wb)
        
        # 2. Vérifier MEMORY_LOG
        self._audit_memory_log(wb)
        
        # 3. Vérifier SNAPSHOT_ACTIVE
        self._audit_snapshot(wb)
        
        # 4. Vérifier RISKS
        self._audit_risks(wb)
        
        # 5. Vérifier CONFLITS_DETECTES
        self._audit_conflicts(wb)
        
        # 6. Vérifier CARTOGRAPHIE_APPELS
        self._audit_cartography(wb)
        
        # 7. Vérifier DEPENDANCES_SCRIPTS
        self._audit_dependencies(wb)
        
        wb.close()
        
        return self.results
        
    def _audit_sheets(self, wb):
        """Audit des onglets du HUB"""
        print("\n📊 Audit onglets HUB...")
        
        expected_sheets = [
            "SETTINGS",
            "MEMORY_LOG",
            "SNAPSHOT_ACTIVE",
            "REGLES_DE_GOUVERNANCE",
            "ARCHITECTURE_GLOBALE",
            "CARTOGRAPHIE_APPELS",
            "DEPENDANCES_SCRIPTS",
            "CONFLITS_DETECTES",
            "RISQUES",
            "LOGS"
        ]
        
        present_sheets = wb.sheetnames
        
        sheets_status = {
            "expected": expected_sheets,
            "present": present_sheets,
            "missing": [s for s in expected_sheets if s not in present_sheets],
            "extra": [s for s in present_sheets if s not in expected_sheets]
        }
        
        self.results["sheets"] = sheets_status
        
        print(f"  ✅ Onglets présents : {len(present_sheets)}")
        print(f"  ✅ Onglets attendus trouvés : {len(expected_sheets) - len(sheets_status['missing'])}/{len(expected_sheets)}")
        
        if sheets_status["missing"]:
            print(f"  ⚠️  Onglets manquants : {sheets_status['missing']}")
            
    def _audit_memory_log(self, wb):
        """Audit de MEMORY_LOG"""
        print("\n📝 Audit MEMORY_LOG...")
        
        if "MEMORY_LOG" not in wb.sheetnames:
            print("  ❌ MEMORY_LOG manquant")
            self.results["memory_log"] = {"error": "sheet_missing"}
            return
            
        ws = wb["MEMORY_LOG"]
        
        # Vérifier structure TSV 7 colonnes
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if max_row > 0 else []
        
        expected_columns = [
            "timestamp",
            "event_type",
            "source",
            "entity_id",
            "action",
            "status",
            "metadata_json"
        ]
        
        memory_log_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col,
            "header": list(header_row) if header_row else [],
            "expected_columns": expected_columns,
            "structure_valid": False
        }
        
        # Vérifier si colonnes correspondent (flexible sur noms exacts)
        if memory_log_status["column_count"] == 7:
            memory_log_status["structure_valid"] = True
            print(f"  ✅ Structure TSV 7 colonnes valide")
        else:
            print(f"  ⚠️  Structure incorrecte : {memory_log_status['column_count']} colonnes (attendu: 7)")
            
        print(f"  ✅ Lignes de log : {memory_log_status['row_count']}")
        
        self.results["memory_log"] = memory_log_status
        
    def _audit_snapshot(self, wb):
        """Audit de SNAPSHOT_ACTIVE"""
        print("\n📸 Audit SNAPSHOT_ACTIVE...")
        
        if "SNAPSHOT_ACTIVE" not in wb.sheetnames:
            print("  ❌ SNAPSHOT_ACTIVE manquant")
            self.results["snapshot"] = {"error": "sheet_missing"}
            return
            
        ws = wb["SNAPSHOT_ACTIVE"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        snapshot_status = {
            "row_count": max_row,
            "column_count": max_col,
            "last_snapshot_date": None
        }
        
        # Chercher date de dernier snapshot
        for row in ws.iter_rows(min_row=1, max_row=min(10, max_row), values_only=True):
            if row and any("snapshot" in str(cell).lower() for cell in row if cell):
                for cell in row:
                    if cell and isinstance(cell, (datetime, str)):
                        snapshot_status["last_snapshot_date"] = str(cell)
                        break
                        
        print(f"  ✅ Lignes snapshot : {snapshot_status['row_count']}")
        if snapshot_status["last_snapshot_date"]:
            print(f"  ✅ Dernier snapshot : {snapshot_status['last_snapshot_date']}")
            
        self.results["snapshot"] = snapshot_status
        
    def _audit_risks(self, wb):
        """Audit de RISQUES"""
        print("\n⚠️  Audit RISQUES...")
        
        if "RISQUES" not in wb.sheetnames:
            print("  ❌ RISQUES manquant")
            self.results["risks"] = {"error": "sheet_missing"}
            return
            
        ws = wb["RISQUES"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        risks_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Risques enregistrés : {risks_status['row_count']}")
        
        self.results["risks"] = risks_status
        
    def _audit_conflicts(self, wb):
        """Audit de CONFLITS_DETECTES"""
        print("\n🔥 Audit CONFLITS_DETECTES...")
        
        if "CONFLITS_DETECTES" not in wb.sheetnames:
            print("  ❌ CONFLITS_DETECTES manquant")
            self.results["conflicts"] = {"error": "sheet_missing"}
            return
            
        ws = wb["CONFLITS_DETECTES"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        conflicts_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Conflits enregistrés : {conflicts_status['row_count']}")
        
        self.results["conflicts"] = conflicts_status
        
    def _audit_cartography(self, wb):
        """Audit de CARTOGRAPHIE_APPELS"""
        print("\n🗺️  Audit CARTOGRAPHIE_APPELS...")
        
        if "CARTOGRAPHIE_APPELS" not in wb.sheetnames:
            print("  ❌ CARTOGRAPHIE_APPELS manquant")
            self.results["cartography"] = {"error": "sheet_missing"}
            return
            
        ws = wb["CARTOGRAPHIE_APPELS"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        cartography_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Appels cartographiés : {cartography_status['row_count']}")
        
        self.results["cartography"] = cartography_status
        
    def _audit_dependencies(self, wb):
        """Audit de DEPENDANCES_SCRIPTS"""
        print("\n🔗 Audit DEPENDANCES_SCRIPTS...")
        
        if "DEPENDANCES_SCRIPTS" not in wb.sheetnames:
            print("  ❌ DEPENDANCES_SCRIPTS manquant")
            self.results["dependencies"] = {"error": "sheet_missing"}
            return
            
        ws = wb["DEPENDANCES_SCRIPTS"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        dependencies_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Dépendances enregistrées : {dependencies_status['row_count']}")
        
        self.results["dependencies"] = dependencies_status


class BOXAudit:
    """Audit de BOX2026 IAPF (Sheets CRM)"""
    
    def __init__(self, box_file: Path):
        self.box_file = box_file
        self.results = {}
        
    def run(self) -> dict:
        """Exécute l'audit BOX2026 complet"""
        print("\n" + "="*80)
        print("🔍 AUDIT BOX2026 IAPF (CRM)")
        print("="*80)
        
        if not self.box_file.exists():
            print(f"  ❌ Fichier BOX non trouvé : {self.box_file}")
            self.results["error"] = "BOX file not found"
            return self.results
            
        # Charger le workbook
        wb = openpyxl.load_workbook(self.box_file, read_only=True)
        
        # 1. Vérifier onglets CRM
        self._audit_crm_sheets(wb)
        
        # 2. Vérifier CONFIG
        self._audit_config(wb)
        
        # 3. Vérifier INDEX_GLOBAL
        self._audit_index(wb)
        
        # 4. Vérifier CRM_DEVIS
        self._audit_devis(wb)
        
        # 5. Vérifier CRM_FACTURES
        self._audit_factures(wb)
        
        wb.close()
        
        return self.results
        
    def _audit_crm_sheets(self, wb):
        """Audit des onglets CRM"""
        print("\n📊 Audit onglets CRM...")
        
        expected_sheets = [
            "CONFIG",
            "INDEX_GLOBAL",
            "LOGS_SYSTEM",
            "COMPTABILITE",
            "CRM_CLIENTS",
            "CRM_DEVIS",
            "CRM_DEVIS_LIGNES",
            "CRM_FACTURES",
            "CRM_EVENTS"
        ]
        
        present_sheets = wb.sheetnames
        
        sheets_status = {
            "expected": expected_sheets,
            "present": present_sheets,
            "missing": [s for s in expected_sheets if s not in present_sheets]
        }
        
        self.results["sheets"] = sheets_status
        
        print(f"  ✅ Onglets CRM présents : {len(expected_sheets) - len(sheets_status['missing'])}/{len(expected_sheets)}")
        
        if sheets_status["missing"]:
            print(f"  ⚠️  Onglets manquants : {sheets_status['missing']}")
            
    def _audit_config(self, wb):
        """Audit de CONFIG"""
        print("\n⚙️  Audit CONFIG...")
        
        if "CONFIG" not in wb.sheetnames:
            print("  ❌ CONFIG manquant")
            self.results["config"] = {"error": "sheet_missing"}
            return
            
        ws = wb["CONFIG"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        config_status = {
            "row_count": max_row,
            "column_count": max_col
        }
        
        print(f"  ✅ Lignes config : {config_status['row_count']}")
        
        self.results["config"] = config_status
        
    def _audit_index(self, wb):
        """Audit de INDEX_GLOBAL"""
        print("\n📇 Audit INDEX_GLOBAL...")
        
        if "INDEX_GLOBAL" not in wb.sheetnames:
            print("  ❌ INDEX_GLOBAL manquant")
            self.results["index"] = {"error": "sheet_missing"}
            return
            
        ws = wb["INDEX_GLOBAL"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        index_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Documents indexés : {index_status['row_count']}")
        
        self.results["index"] = index_status
        
    def _audit_devis(self, wb):
        """Audit de CRM_DEVIS"""
        print("\n📋 Audit CRM_DEVIS...")
        
        if "CRM_DEVIS" not in wb.sheetnames:
            print("  ❌ CRM_DEVIS manquant")
            self.results["devis"] = {"error": "sheet_missing"}
            return
            
        ws = wb["CRM_DEVIS"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        devis_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Devis enregistrés : {devis_status['row_count']}")
        
        self.results["devis"] = devis_status
        
    def _audit_factures(self, wb):
        """Audit de CRM_FACTURES"""
        print("\n💰 Audit CRM_FACTURES...")
        
        if "CRM_FACTURES" not in wb.sheetnames:
            print("  ❌ CRM_FACTURES manquant")
            self.results["factures"] = {"error": "sheet_missing"}
            return
            
        ws = wb["CRM_FACTURES"]
        
        max_row = ws.max_row if ws.max_row is not None else 0
        max_col = ws.max_column if ws.max_column is not None else 0
        
        factures_status = {
            "row_count": max_row - 1 if max_row > 0 else 0,
            "column_count": max_col
        }
        
        print(f"  ✅ Factures enregistrées : {factures_status['row_count']}")
        
        self.results["factures"] = factures_status


def main():
    """Point d'entrée principal"""
    print("\n" + "="*80)
    print("🚀 AUDIT GLOBAL SYSTÈME IAPF - MODE PROPOSAL-FIRST")
    print("="*80)
    print(f"📅 Date : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📁 Répertoire : {WEBAPP_DIR}")
    print("="*80)
    
    # Initialiser le rapport
    report = AuditReport()
    
    # PHASE 1.1 - Audit OCR (Repo 1)
    ocr_audit = OCRAudit(WEBAPP_DIR)
    ocr_results = ocr_audit.run()
    report.add_section("ocr_repo1", ocr_results)
    
    # PHASE 1.3 - Audit HUB
    hub_audit = HUBAudit(HUB_FILE)
    hub_results = hub_audit.run()
    report.add_section("hub_memory", hub_results)
    
    # PHASE 1.2 - Audit BOX2026
    box_audit = BOXAudit(BOX_FILE)
    box_results = box_audit.run()
    report.add_section("box2026_crm", box_results)
    
    # Ajouter des propositions basées sur les audits
    print("\n" + "="*80)
    print("💡 GÉNÉRATION DES PROPOSITIONS")
    print("="*80)
    
    # Propositions OCR
    if not ocr_results.get("governance", {}).get("read_only_enforced"):
        report.add_proposal(
            "OCR",
            "Vérifier enforcement du mode READ-ONLY dans Cloud Run",
            "high"
        )
        
    # Propositions HUB
    if hub_results.get("memory_log", {}).get("row_count", 0) == 0:
        report.add_proposal(
            "HUB",
            "MEMORY_LOG vide - initialiser avec événements système",
            "medium"
        )
        
    # Propositions CRM
    if box_results.get("devis", {}).get("row_count", 0) > 0:
        report.add_proposal(
            "CRM",
            "Analyser pipeline devis → facture pour cohérence numérotation",
            "high"
        )
        
    # Sauvegarder le rapport
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = WEBAPP_DIR / f"audit_global_iapf_{timestamp}.json"
    report.save(output_file)
    
    # Résumé final
    print("\n" + "="*80)
    print("✅ AUDIT GLOBAL TERMINÉ")
    print("="*80)
    print(f"📊 Sections auditées : {len(report.sections)}")
    print(f"💡 Propositions générées : {len(report.proposals)}")
    print(f"🔧 Corrections identifiées : {len(report.corrections)}")
    print(f"⚠️  Risques détectés : {len(report.risks)}")
    print(f"🔥 Conflits détectés : {len(report.conflicts)}")
    print("="*80)
    
    return report


if __name__ == "__main__":
    report = main()
